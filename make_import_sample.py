"""Génère sample_import.xlsx — fichier de test pour la fonction Import."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "sample_import.xlsx")

HEADERS = [
    "Nom et Prénom",
    "Statut",
    "Filiale du Groupe",
    "Description",
    "Service",
    "Société",
    "Rôle",
    "Domaine",
    "Date d'octroi",
    "Date des attestations",
]

ROWS = [
    ("THOMAS Élodie",      "Actif",    "Meridia France",        "Accès aux systèmes d'information",     "DSI",                    "Meridia SA",             "Utilisateur",    "Informatique", "15/01/2025", "15/01/2027"),
    ("LEGRAND Sébastien",  "Actif",    "Meridia Digital",       "Développement applicatif",              "DSI",                    "Meridia Digital SAS",    "Administrateur", "Informatique", "03/06/2024", "03/06/2026"),
    ("HAMON Claire",       "Actif",    "Meridia Services",      "Consultation des données financières",  "Direction Financière",   "Meridia SA",             "Lecteur",        "Finance",      "20/09/2024", "20/09/2026"),
    ("VIDAL Arnaud",       "Suspendu", "Meridia Industrie",     "Gestion des accès utilisateurs",       "Ressources Humaines",    "Meridia Industrie SARL", "Utilisateur",    "RH",           "01/03/2023", "01/03/2025"),
    ("CHEVALIER Mélanie",  "Actif",    "Meridia International", "Audit et contrôle interne",             "Juridique & Conformité", "MeriTech SAS",           "Lecteur",        "Juridique",    "12/11/2024", "12/11/2026"),
    ("PICARD Jérémy",      "Actif",    "Meridia France",        "Administration des bases de données",   "DSI",                    "Meridia SA",             "Administrateur", "Informatique", "07/02/2025", "07/02/2027"),
    ("FERNANDEZ Sofia",    "Révoqué",  "Meridia Services",      "Accès aux systèmes d'information",     "Commercial",             "Meridia SA",             "Lecteur",        "Informatique", "15/04/2022", "15/04/2024"),
    ("GUERIN Baptiste",    "Actif",    "Meridia Digital",       "Gestion des accès utilisateurs",       "Ressources Humaines",    "Meridia Digital SAS",    "Utilisateur",    "RH",           "30/08/2024", "30/08/2026"),
    ("MULLER Stéphanie",   "Actif",    "Meridia France",        "Consultation des données financières",  "Direction Financière",   "Meridia SA",             "Administrateur", "Finance",      "18/05/2024", "18/05/2026"),
    ("BOYER Théo",         "Actif",    "Meridia Industrie",     "Supervision des opérations industrielles", "Production",         "Meridia Industrie SARL", "Utilisateur",    "Production",   "22/10/2024", "22/10/2026"),
]

wb = Workbook()
ws = wb.active
ws.title = "Habilitations"

# ── Style en-têtes ────────────────────────────────────────────────────────────
header_fill = PatternFill("solid", fgColor="0D9488")   # teal-600
header_font = Font(bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

# ── Données ───────────────────────────────────────────────────────────────────
alt_fill   = PatternFill("solid", fgColor="F0FDFA")   # teal-50
white_fill = PatternFill("solid", fgColor="FFFFFF")
data_align = Alignment(vertical="center")

STATUS_COLORS = {
    "Actif":    "166534",   # green-800
    "Suspendu": "854D0E",   # yellow-800
    "Révoqué":  "991B1B",   # red-800
}

for row_idx, row in enumerate(ROWS, 2):
    fill = alt_fill if row_idx % 2 == 0 else white_fill
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.alignment = data_align
        cell.border = border
        # Colorer la colonne Statut
        if col_idx == 2 and val in STATUS_COLORS:
            cell.font = Font(bold=True, color=STATUS_COLORS[val])

# ── Largeurs de colonnes ──────────────────────────────────────────────────────
col_widths = [24, 11, 24, 42, 26, 24, 15, 14, 14, 18]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.row_dimensions[1].height = 30

# ── Figer la ligne d'en-tête ──────────────────────────────────────────────────
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"Fichier cree : {OUT}")
print(f"{len(ROWS)} lignes, {len(HEADERS)} colonnes.")
